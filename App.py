import customtkinter as ctk
import mysql.connector
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib
import csv
import openpyxl
matplotlib.use("TkAgg")

# ─────────────────────────────────────────
#  THEME
# ─────────────────────────────────────────
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

NAVY    = "#0a1628"
OCEAN   = "#1a3a5c"
TEAL    = "#0e7490"
SEAFOAM = "#22d3ee"
WHITE   = "#f0f9ff"
GRAY    = "#94a3b8"
RED     = "#ef4444"
GREEN   = "#22c55e"
GOLD    = "#f59e0b"

CATEGORIES = ["Drinks", "Dining", "Excursions", "Entertainment", "Spa"]

# ─────────────────────────────────────────
#  DATABASE
# ─────────────────────────────────────────
def get_connection():
    return mysql.connector.connect(
        host="127.0.0.1", user="root", password="root1234", database="CruiseDB"
    )

def setup_new_tables():
    conn = get_connection(); cursor = conn.cursor()
    cursor.execute("""CREATE TABLE IF NOT EXISTS Products (
                                                              product_id INT AUTO_INCREMENT PRIMARY KEY,
                                                              name VARCHAR(150) NOT NULL,
        category VARCHAR(50) NOT NULL,
        price DECIMAL(10,2) NOT NULL)""")
    cursor.execute("""CREATE TABLE IF NOT EXISTS Purchases (
                                                               purchase_id INT AUTO_INCREMENT PRIMARY KEY,
                                                               passenger_id INT NOT NULL,
                                                               product_id INT NOT NULL,
                                                               quantity INT NOT NULL DEFAULT 1,
                                                               purchase_date DATE NOT NULL,
                                                               FOREIGN KEY (passenger_id) REFERENCES Passengers(passenger_id),
        FOREIGN KEY (product_id) REFERENCES Products(product_id))""")
    conn.commit(); cursor.close(); conn.close()

def cruise_exists():
    conn = get_connection(); cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM Cruises")
    count = cursor.fetchone()[0]
    cursor.close(); conn.close()
    return count > 0

# ─────────────────────────────────────────
#  VALIDATION
# ─────────────────────────────────────────
def validate_text(value, field_name, min_len=2, max_len=150):
    value = value.strip()
    if not value: return False, f"{field_name} cannot be empty."
    if len(value) < min_len: return False, f"{field_name} must be at least {min_len} characters."
    if len(value) > max_len: return False, f"{field_name} cannot exceed {max_len} characters."
    return True, value

def validate_date(value, field_name):
    value = value.strip()
    if not value: return False, f"{field_name} cannot be empty."
    try:
        parsed = datetime.strptime(value, "%Y-%m-%d")
        if parsed.year < 2000 or parsed.year > 2100:
            return False, f"{field_name} year must be between 2000 and 2100."
        return True, value
    except ValueError:
        return False, f"{field_name} must be YYYY-MM-DD (e.g. 2025-08-15)."

def validate_int(value, field_name, min_val=1, max_val=99999):
    value = str(value).strip()
    if not value: return False, f"{field_name} cannot be empty."
    if not value.isdigit(): return False, f"{field_name} must be a whole number."
    value = int(value)
    if value < min_val: return False, f"{field_name} must be at least {min_val}."
    if value > max_val: return False, f"{field_name} cannot exceed {max_val}."
    return True, value

def validate_price(value):
    value = str(value).strip()
    if not value: return False, "Price cannot be empty."
    try:
        v = float(value)
        if v < 0: return False, "Price cannot be negative."
        if v > 99999: return False, "Price max is $99,999."
        return True, round(v, 2)
    except ValueError:
        return False, "Price must be a number (e.g. 12.99)."

def validate_age(value):
    value = str(value).strip()
    if not value or not value.isdigit(): return False, "Age must be a whole number."
    v = int(value)
    if v < 1 or v > 120: return False, "Age must be between 1 and 120."
    return True, v

def validate_cabin(value):
    value = str(value).strip().upper()
    if not value or len(value) < 2 or len(value) > 10:
        return False, "Cabin must be 2-10 characters (e.g. A101)."
    if not value[0].isalpha() or not value[1:].isdigit():
        return False, "Cabin must start with a letter then numbers (e.g. A101)."
    return True, value

# ─────────────────────────────────────────
#  FILE READER
# ─────────────────────────────────────────
def read_file(filepath):
    rows = []
    if filepath.endswith(".csv"):
        with open(filepath, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append({k.strip().lower(): v.strip() for k, v in row.items()})
    elif filepath.endswith(".xlsx"):
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers = [str(cell.value).strip().lower() for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                rows.append({headers[i]: str(row[i]).strip() if row[i] is not None else "" for i in range(len(headers))})
    return rows

# ─────────────────────────────────────────
#  REUSABLE WIDGETS
# ─────────────────────────────────────────
class FormDialog(ctk.CTkToplevel):
    def __init__(self, parent, title, fields, on_submit, prefill=None):
        super().__init__(parent)
        self.title(title)
        self.geometry("480x" + str(min(140 + len(fields) * 82, 640)))
        self.resizable(False, False)
        self.configure(fg_color=NAVY)
        self.grab_set()
        self.entries = {}
        self.on_submit = on_submit

        ctk.CTkLabel(self, text=title, font=("Helvetica", 18, "bold"),
                     text_color=SEAFOAM).pack(pady=(24,12), padx=24)

        sf = ctk.CTkScrollableFrame(self, fg_color=NAVY)
        sf.pack(fill="both", expand=True, padx=24)

        for field_info in fields:
            field_name  = field_info[0]
            placeholder = field_info[1]
            opts        = field_info[2] if len(field_info) > 2 else None
            ctk.CTkLabel(sf, text=field_name, font=("Helvetica", 13),
                         text_color=WHITE, anchor="w").pack(pady=(8,2), fill="x")
            if opts and isinstance(opts, list):
                var = tk.StringVar(value=opts[0])
                ctk.CTkOptionMenu(sf, values=opts, variable=var,
                                  fg_color=OCEAN, button_color=TEAL,
                                  text_color=WHITE, height=38).pack(fill="x")
                if prefill and field_name in prefill:
                    var.set(prefill[field_name])
                self.entries[field_name] = var
            else:
                entry = ctk.CTkEntry(sf, placeholder_text=placeholder,
                                     fg_color=OCEAN, border_color=TEAL,
                                     text_color=WHITE, height=38)
                entry.pack(fill="x")
                if prefill and field_name in prefill:
                    entry.insert(0, prefill[field_name])
                self.entries[field_name] = entry

        self.error_label = ctk.CTkLabel(self, text="", font=("Helvetica", 12),
                                        text_color=RED, wraplength=420)
        self.error_label.pack(pady=(8,0), padx=24)

        bf = ctk.CTkFrame(self, fg_color="transparent")
        bf.pack(pady=16, padx=24, fill="x")
        ctk.CTkButton(bf, text="❌  Cancel", fg_color=OCEAN, hover_color=TEAL,
                      text_color=WHITE, command=self.destroy).pack(side="left", expand=True, padx=(0,6))
        ctk.CTkButton(bf, text="💾  Save", fg_color=TEAL, hover_color=SEAFOAM,
                      text_color=NAVY, font=("Helvetica", 13, "bold"),
                      command=self._submit).pack(side="left", expand=True, padx=(6,0))

    def _submit(self):
        values = {k: w.get() for k, w in self.entries.items()}
        if self.on_submit(values, self.error_label):
            self.destroy()


class DataTable(ctk.CTkFrame):
    def __init__(self, parent, columns, **kwargs):
        super().__init__(parent, fg_color=OCEAN, corner_radius=10, **kwargs)
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("CT.Treeview", background=OCEAN, foreground=WHITE,
                        fieldbackground=OCEAN, rowheight=30, font=("Helvetica", 12))
        style.configure("CT.Treeview.Heading", background=TEAL, foreground=WHITE,
                        font=("Helvetica", 12, "bold"), relief="flat")
        style.map("CT.Treeview", background=[("selected", TEAL)])
        self.tree = ttk.Treeview(self, columns=columns, show="headings", style="CT.Treeview")
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=max(80, int(950/len(columns))), anchor="center")
        sb = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side="left", fill="both", expand=True, padx=4, pady=4)
        sb.pack(side="right", fill="y", pady=4)

    def load(self, rows):
        for i in self.tree.get_children(): self.tree.delete(i)
        for row in rows: self.tree.insert("", "end", values=row)

    def selected_id(self):
        sel = self.tree.selection()
        return self.tree.item(sel[0])["values"][0] if sel else None

    def selected_row(self):
        sel = self.tree.selection()
        return self.tree.item(sel[0])["values"] if sel else None


def tab_header(parent, title, btns, refresh_cmd):
    hdr = ctk.CTkFrame(parent, fg_color="transparent")
    hdr.pack(fill="x", padx=20, pady=(20,10))
    ctk.CTkLabel(hdr, text=title, font=("Helvetica", 22, "bold"),
                 text_color=SEAFOAM).pack(side="left")
    bf = ctk.CTkFrame(hdr, fg_color="transparent")
    bf.pack(side="right")
    for label, fc, hc, tc, cmd in btns:
        ctk.CTkButton(bf, text=label, fg_color=fc, hover_color=hc,
                      text_color=tc, width=110, command=cmd).pack(side="left", padx=3)
    ctk.CTkButton(bf, text="↻  Refresh", fg_color=OCEAN, hover_color=TEAL,
                  text_color=WHITE, width=100, command=refresh_cmd).pack(side="left", padx=3)


def upload_btn(parent, label, cmd):
    ctk.CTkButton(parent, text=label, fg_color=GOLD, hover_color="#d97706",
                  text_color=NAVY, width=130, font=("Helvetica", 12, "bold"),
                  command=cmd).pack(side="left", padx=3)


# ─────────────────────────────────────────
#  GATE SCREEN
# ─────────────────────────────────────────
class GateScreen(ctk.CTkFrame):
    def __init__(self, parent, on_unlock):
        super().__init__(parent, fg_color=NAVY)
        self.on_unlock = on_unlock
        self.pack(fill="both", expand=True)

        ctk.CTkLabel(self, text="🚢", font=("Helvetica", 72)).pack(pady=(80,10))
        ctk.CTkLabel(self, text="Welcome to Cruise Trip Tracker",
                     font=("Helvetica", 26, "bold"), text_color=SEAFOAM).pack()
        ctk.CTkLabel(self, text="You must add at least one cruise before anything else.",
                     font=("Helvetica", 14), text_color=GRAY).pack(pady=(8,40))

        fields_frame = ctk.CTkFrame(self, fg_color=OCEAN, corner_radius=12)
        fields_frame.pack(padx=200)

        ctk.CTkLabel(fields_frame, text="Add Your First Cruise",
                     font=("Helvetica", 16, "bold"), text_color=WHITE).pack(pady=(20,12), padx=30)

        self.entries = {}
        form_fields = [
            ("Cruise Line",     "e.g. Royal Caribbean"),
            ("Ship Name",       "e.g. Wonder of the Seas"),
            ("Destination",     "e.g. Caribbean"),
            ("Departure Date",  "YYYY-MM-DD"),
            ("Duration (days)", "e.g. 7"),
        ]
        for fname, ph in form_fields:
            ctk.CTkLabel(fields_frame, text=fname, font=("Helvetica", 12),
                         text_color=WHITE, anchor="w").pack(fill="x", padx=24, pady=(6,1))
            e = ctk.CTkEntry(fields_frame, placeholder_text=ph, fg_color=NAVY,
                             border_color=TEAL, text_color=WHITE, height=36)
            e.pack(fill="x", padx=24)
            self.entries[fname] = e

        self.err = ctk.CTkLabel(fields_frame, text="", font=("Helvetica", 12),
                                text_color=RED, wraplength=380)
        self.err.pack(pady=(8,0))

        ctk.CTkButton(fields_frame, text="🚀  Launch App", fg_color=TEAL, hover_color=SEAFOAM,
                      text_color=NAVY, font=("Helvetica", 14, "bold"), height=42,
                      command=self._submit).pack(pady=20, padx=24, fill="x")

    def _submit(self):
        v = {k: e.get() for k, e in self.entries.items()}
        ok, cl   = validate_text(v["Cruise Line"],    "Cruise Line");
        if not ok: self.err.configure(text=cl);   return
        ok, sn   = validate_text(v["Ship Name"],      "Ship Name");
        if not ok: self.err.configure(text=sn);   return
        ok, dest = validate_text(v["Destination"],    "Destination");
        if not ok: self.err.configure(text=dest); return
        ok, dep  = validate_date(v["Departure Date"], "Departure Date");
        if not ok: self.err.configure(text=dep);  return
        ok, dur  = validate_int(v["Duration (days)"], "Duration", 1, 365)
        if not ok: self.err.configure(text=dur);  return
        conn = get_connection(); cursor = conn.cursor()
        cursor.execute("INSERT INTO Cruises (cruise_line,ship_name,destination,departure_date,duration_days) VALUES (%s,%s,%s,%s,%s)",
                       (cl, sn, dest, dep, dur))
        conn.commit(); cursor.close(); conn.close()
        self.destroy()
        self.on_unlock()


# ─────────────────────────────────────────
#  UPLOAD RESULT DIALOG
# ─────────────────────────────────────────
class UploadResultDialog(ctk.CTkToplevel):
    def __init__(self, parent, title, success, errors):
        super().__init__(parent)
        self.title(title)
        self.geometry("500x420")
        self.configure(fg_color=NAVY)
        self.grab_set()

        ctk.CTkLabel(self, text=title, font=("Helvetica", 16, "bold"),
                     text_color=SEAFOAM).pack(pady=(20,10))
        ctk.CTkLabel(self, text=f"✅  {success} rows imported successfully",
                     font=("Helvetica", 13), text_color=GREEN).pack()
        if errors:
            ctk.CTkLabel(self, text=f"⚠️  {len(errors)} rows had errors:",
                         font=("Helvetica", 13), text_color=GOLD).pack(pady=(8,4))
            sf = ctk.CTkScrollableFrame(self, fg_color=OCEAN, height=200)
            sf.pack(fill="both", expand=True, padx=20, pady=(0,10))
            for e in errors:
                ctk.CTkLabel(sf, text=e, font=("Helvetica", 11),
                             text_color=RED, anchor="w", wraplength=440).pack(fill="x", pady=2)
        ctk.CTkButton(self, text="✅  Close", fg_color=TEAL, hover_color=SEAFOAM,
                      text_color=NAVY, command=self.destroy).pack(pady=12)


# ─────────────────────────────────────────
#  TEMPLATE DOWNLOAD HELPER
# ─────────────────────────────────────────
def download_template(headers, default_name):
    path = filedialog.asksaveasfilename(
        defaultextension=".csv",
        filetypes=[("CSV files","*.csv")],
        initialfile=default_name
    )
    if path:
        with open(path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
        messagebox.showinfo("Template Saved", f"Template saved to:\n{path}")


# ─────────────────────────────────────────
#  CRUISES TAB
# ─────────────────────────────────────────
class CruisesTab(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent, fg_color=NAVY)
        tab_header(self, "🚢  Cruises", [
            ("➕  Add",    TEAL, SEAFOAM, NAVY, self._add),
            ("✏️  Edit",   OCEAN, TEAL, WHITE,  self._edit),
            ("🗑️  Delete", RED, "#b91c1c", WHITE, self._delete),
        ], self.refresh)

        uf = ctk.CTkFrame(self, fg_color="transparent")
        uf.pack(fill="x", padx=20, pady=(0,8))
        upload_btn(uf, "📂  Upload CSV/Excel", self._upload)
        ctk.CTkButton(uf, text="📋  Download Template", fg_color=OCEAN, hover_color=TEAL,
                      text_color=WHITE, width=160,
                      command=lambda: download_template(
                          ["cruise_line","ship_name","destination","departure_date","duration_days"],
                          "cruises_template.csv"
                      )).pack(side="left", padx=3)

        self.table = DataTable(self, ["ID","Cruise Line","Ship Name","Destination","Departure","Days"])
        self.table.pack(fill="both", expand=True, padx=20, pady=(0,20))
        self.refresh()

    def refresh(self):
        conn = get_connection(); cursor = conn.cursor()
        cursor.execute("SELECT * FROM Cruises")
        self.table.load(cursor.fetchall())
        cursor.close(); conn.close()

    def _add(self):
        fields = [
            ("Cruise Line",    "e.g. Royal Caribbean", None),
            ("Ship Name",      "e.g. Wonder of the Seas", None),
            ("Destination",    "e.g. Caribbean", None),
            ("Departure Date", "YYYY-MM-DD", None),
            ("Duration (days)","e.g. 7", None),
        ]
        def submit(v, err):
            ok,cl   = validate_text(v["Cruise Line"],    "Cruise Line");
            if not ok: err.configure(text=cl);   return False
            ok,sn   = validate_text(v["Ship Name"],      "Ship Name");
            if not ok: err.configure(text=sn);   return False
            ok,dest = validate_text(v["Destination"],    "Destination");
            if not ok: err.configure(text=dest); return False
            ok,dep  = validate_date(v["Departure Date"], "Departure Date");
            if not ok: err.configure(text=dep);  return False
            ok,dur  = validate_int(v["Duration (days)"], "Duration", 1, 365)
            if not ok: err.configure(text=dur);  return False
            conn = get_connection(); cursor = conn.cursor()
            cursor.execute("INSERT INTO Cruises (cruise_line,ship_name,destination,departure_date,duration_days) VALUES (%s,%s,%s,%s,%s)",
                           (cl,sn,dest,dep,dur))
            conn.commit(); cursor.close(); conn.close(); self.refresh(); return True
        FormDialog(self, "➕  Add Cruise", fields, submit)

    def _edit(self):
        row = self.table.selected_row()
        if not row: messagebox.showwarning("⚠️ No Selection","Please select a cruise first."); return
        fields = [("Cruise Line","",None),("Ship Name","",None),("Destination","",None),
                  ("Departure Date","YYYY-MM-DD",None),("Duration (days)","",None)]
        prefill = {"Cruise Line":str(row[1]),"Ship Name":str(row[2]),
                   "Destination":str(row[3]),"Departure Date":str(row[4]),"Duration (days)":str(row[5])}
        cid = row[0]
        def submit(v, err):
            ok,cl   = validate_text(v["Cruise Line"],    "Cruise Line");
            if not ok: err.configure(text=cl);   return False
            ok,sn   = validate_text(v["Ship Name"],      "Ship Name");
            if not ok: err.configure(text=sn);   return False
            ok,dest = validate_text(v["Destination"],    "Destination");
            if not ok: err.configure(text=dest); return False
            ok,dep  = validate_date(v["Departure Date"], "Departure Date");
            if not ok: err.configure(text=dep);  return False
            ok,dur  = validate_int(v["Duration (days)"], "Duration", 1, 365)
            if not ok: err.configure(text=dur);  return False
            conn = get_connection(); cursor = conn.cursor()
            cursor.execute("UPDATE Cruises SET cruise_line=%s,ship_name=%s,destination=%s,departure_date=%s,duration_days=%s WHERE cruise_id=%s",
                           (cl,sn,dest,dep,dur,cid))
            conn.commit(); cursor.close(); conn.close(); self.refresh(); return True
        FormDialog(self, "✏️  Edit Cruise", fields, submit, prefill)

    def _delete(self):
        cid = self.table.selected_id()
        if not cid: messagebox.showwarning("⚠️ No Selection","Please select a cruise first."); return
        conn = get_connection(); cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM Passengers WHERE cruise_id=%s",(cid,)); pc=cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM Ports WHERE cruise_id=%s",(cid,)); portc=cursor.fetchone()[0]
        if pc>0 or portc>0:
            messagebox.showerror("🚫 Cannot Delete",f"Cruise {cid} has {pc} passenger(s) and {portc} port(s).\nDelete those first.")
            cursor.close(); conn.close(); return
        if messagebox.askyesno("🗑️ Confirm Delete",f"Delete Cruise {cid}? This cannot be undone."):
            cursor.execute("DELETE FROM Cruises WHERE cruise_id=%s",(cid,)); conn.commit()
        cursor.close(); conn.close(); self.refresh()

    def _upload(self):
        path = filedialog.askopenfilename(filetypes=[("CSV/Excel","*.csv *.xlsx")])
        if not path: return
        try: rows = read_file(path)
        except Exception as e: messagebox.showerror("❌ File Error", str(e)); return
        required = {"cruise_line","ship_name","destination","departure_date","duration_days"}
        if not required.issubset(set(rows[0].keys()) if rows else set()):
            messagebox.showerror("❌ Wrong Format",
                                 f"File must have columns:\n{', '.join(required)}\n\nDownload the template to get started.")
            return
        success=0; errors=[]
        conn = get_connection(); cursor = conn.cursor()
        for i, row in enumerate(rows, 2):
            try:
                ok,cl   = validate_text(row.get("cruise_line",""),   "cruise_line")
                ok2,sn  = validate_text(row.get("ship_name",""),     "ship_name")
                ok3,dst = validate_text(row.get("destination",""),   "destination")
                ok4,dep = validate_date(row.get("departure_date",""),"departure_date")
                ok5,dur = validate_int(row.get("duration_days",""),  "duration_days", 1, 365)
                if not all([ok,ok2,ok3,ok4,ok5]):
                    errors.append(f"Row {i}: {[cl,sn,dst,dep,dur][([ok,ok2,ok3,ok4,ok5]).index(False)]}"); continue
                cursor.execute("INSERT INTO Cruises (cruise_line,ship_name,destination,departure_date,duration_days) VALUES (%s,%s,%s,%s,%s)",
                               (cl,sn,dst,dep,dur)); success+=1
            except Exception as e: errors.append(f"Row {i}: {e}")
        conn.commit(); cursor.close(); conn.close()
        self.refresh()
        UploadResultDialog(self, "📂 Cruise Upload Results", success, errors)


# ─────────────────────────────────────────
#  PASSENGERS TAB
# ─────────────────────────────────────────
class PassengersTab(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent, fg_color=NAVY)
        tab_header(self, "👤  Passengers", [
            ("➕  Add",    TEAL, SEAFOAM, NAVY, self._add),
            ("✏️  Edit",   OCEAN, TEAL, WHITE,  self._edit),
            ("🗑️  Delete", RED, "#b91c1c", WHITE, self._delete),
        ], self.refresh)

        uf = ctk.CTkFrame(self, fg_color="transparent")
        uf.pack(fill="x", padx=20, pady=(0,8))
        upload_btn(uf, "📂  Upload CSV/Excel", self._upload)
        ctk.CTkButton(uf, text="📋  Download Template", fg_color=OCEAN, hover_color=TEAL,
                      text_color=WHITE, width=160,
                      command=lambda: download_template(
                          ["first_name","last_name","age","cabin_number","cruise_id"],
                          "passengers_template.csv"
                      )).pack(side="left", padx=3)

        self.table = DataTable(self, ["ID","First Name","Last Name","Age","Cabin","Cruise ID"])
        self.table.pack(fill="both", expand=True, padx=20, pady=(0,20))
        self.refresh()

    def refresh(self):
        conn = get_connection(); cursor = conn.cursor()
        cursor.execute("SELECT * FROM Passengers")
        self.table.load(cursor.fetchall())
        cursor.close(); conn.close()

    def _cruise_ids(self):
        conn = get_connection(); cursor = conn.cursor()
        cursor.execute("SELECT cruise_id FROM Cruises")
        ids = [str(r[0]) for r in cursor.fetchall()]
        cursor.close(); conn.close(); return ids

    def _add(self):
        ids = self._cruise_ids()
        fields = [
            ("First Name",   "e.g. LeBron", None),
            ("Last Name",    "e.g. James",  None),
            ("Age",          "e.g. 39",     None),
            ("Cabin Number", "e.g. A101",   None),
            ("Cruise ID",    f"One of: {', '.join(ids)}", None),
        ]
        def submit(v, err):
            ok,fn  = validate_text(v["First Name"], "First Name");
            if not ok: err.configure(text=fn);  return False
            ok,ln  = validate_text(v["Last Name"],  "Last Name");
            if not ok: err.configure(text=ln);  return False
            ok,age = validate_age(v["Age"]);
            if not ok: err.configure(text=age); return False
            ok,cab = validate_cabin(v["Cabin Number"]);
            if not ok: err.configure(text=cab); return False
            ok,cid = validate_int(v["Cruise ID"],"Cruise ID");
            if not ok: err.configure(text=cid); return False
            conn = get_connection(); cursor = conn.cursor()
            cursor.execute("SELECT cruise_id FROM Cruises WHERE cruise_id=%s",(cid,))
            if not cursor.fetchone(): err.configure(text=f"Cruise ID {cid} not found."); cursor.close(); conn.close(); return False
            cursor.execute("SELECT passenger_id FROM Passengers WHERE cabin_number=%s AND cruise_id=%s",(cab,cid))
            if cursor.fetchone(): err.configure(text=f"Cabin {cab} already taken on Cruise {cid}."); cursor.close(); conn.close(); return False
            cursor.execute("INSERT INTO Passengers (first_name,last_name,age,cabin_number,cruise_id) VALUES (%s,%s,%s,%s,%s)",
                           (fn,ln,age,cab,cid))
            conn.commit(); cursor.close(); conn.close(); self.refresh(); return True
        FormDialog(self, "➕  Add Passenger", fields, submit)

    def _edit(self):
        row = self.table.selected_row()
        if not row: messagebox.showwarning("⚠️ No Selection","Please select a passenger first."); return
        fields = [("First Name","",None),("Last Name","",None),("Age","",None),
                  ("Cabin Number","",None),("Cruise ID","",None)]
        prefill = {"First Name":str(row[1]),"Last Name":str(row[2]),
                   "Age":str(row[3]),"Cabin Number":str(row[4]),"Cruise ID":str(row[5])}
        pid = row[0]
        def submit(v, err):
            ok,fn  = validate_text(v["First Name"], "First Name");
            if not ok: err.configure(text=fn);  return False
            ok,ln  = validate_text(v["Last Name"],  "Last Name");
            if not ok: err.configure(text=ln);  return False
            ok,age = validate_age(v["Age"]);
            if not ok: err.configure(text=age); return False
            ok,cab = validate_cabin(v["Cabin Number"]);
            if not ok: err.configure(text=cab); return False
            ok,cid = validate_int(v["Cruise ID"],"Cruise ID");
            if not ok: err.configure(text=cid); return False
            conn = get_connection(); cursor = conn.cursor()
            cursor.execute("SELECT passenger_id FROM Passengers WHERE cabin_number=%s AND cruise_id=%s AND passenger_id!=%s",(cab,cid,pid))
            if cursor.fetchone(): err.configure(text=f"Cabin {cab} already taken."); cursor.close(); conn.close(); return False
            cursor.execute("UPDATE Passengers SET first_name=%s,last_name=%s,age=%s,cabin_number=%s,cruise_id=%s WHERE passenger_id=%s",
                           (fn,ln,age,cab,cid,pid))
            conn.commit(); cursor.close(); conn.close(); self.refresh(); return True
        FormDialog(self, "✏️  Edit Passenger", fields, submit, prefill)

    def _delete(self):
        pid = self.table.selected_id()
        if not pid: messagebox.showwarning("⚠️ No Selection","Please select a passenger first."); return
        conn = get_connection(); cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM Purchases WHERE passenger_id=%s",(pid,))
        pc = cursor.fetchone()[0]
        if pc > 0:
            messagebox.showerror("🚫 Cannot Delete",f"Passenger {pid} has {pc} purchase(s).\nDelete purchases first.")
            cursor.close(); conn.close(); return
        if messagebox.askyesno("🗑️ Confirm Delete",f"Delete Passenger {pid}?"):
            cursor.execute("DELETE FROM Passengers WHERE passenger_id=%s",(pid,)); conn.commit()
        cursor.close(); conn.close(); self.refresh()

    def _upload(self):
        path = filedialog.askopenfilename(filetypes=[("CSV/Excel","*.csv *.xlsx")])
        if not path: return
        try: rows = read_file(path)
        except Exception as e: messagebox.showerror("❌ File Error", str(e)); return
        required = {"first_name","last_name","age","cabin_number","cruise_id"}
        if not required.issubset(set(rows[0].keys()) if rows else set()):
            messagebox.showerror("❌ Wrong Format",f"File must have columns:\n{', '.join(required)}"); return
        success=0; errors=[]
        conn = get_connection(); cursor = conn.cursor()
        for i, row in enumerate(rows, 2):
            try:
                ok,fn  = validate_text(row.get("first_name",""),  "first_name")
                ok2,ln = validate_text(row.get("last_name",""),   "last_name")
                ok3,age= validate_age(row.get("age",""))
                ok4,cab= validate_cabin(row.get("cabin_number",""))
                ok5,cid= validate_int(row.get("cruise_id",""),    "cruise_id")
                if not all([ok,ok2,ok3,ok4,ok5]):
                    errors.append(f"Row {i}: validation failed"); continue
                cursor.execute("SELECT cruise_id FROM Cruises WHERE cruise_id=%s",(cid,))
                if not cursor.fetchone(): errors.append(f"Row {i}: Cruise ID {cid} not found"); continue
                cursor.execute("SELECT passenger_id FROM Passengers WHERE cabin_number=%s AND cruise_id=%s",(cab,cid))
                if cursor.fetchone(): errors.append(f"Row {i}: Cabin {cab} already taken on Cruise {cid}"); continue
                cursor.execute("INSERT INTO Passengers (first_name,last_name,age,cabin_number,cruise_id) VALUES (%s,%s,%s,%s,%s)",
                               (fn,ln,age,cab,cid)); success+=1
            except Exception as e: errors.append(f"Row {i}: {e}")
        conn.commit(); cursor.close(); conn.close()
        self.refresh()
        UploadResultDialog(self, "📂 Passenger Upload Results", success, errors)


# ─────────────────────────────────────────
#  PORTS TAB
# ─────────────────────────────────────────
class PortsTab(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent, fg_color=NAVY)
        tab_header(self, "⚓  Ports", [
            ("➕  Add",    TEAL, SEAFOAM, NAVY, self._add),
            ("✏️  Edit",   OCEAN, TEAL, WHITE,  self._edit),
            ("🗑️  Delete", RED, "#b91c1c", WHITE, self._delete),
        ], self.refresh)

        uf = ctk.CTkFrame(self, fg_color="transparent")
        uf.pack(fill="x", padx=20, pady=(0,8))
        upload_btn(uf, "📂  Upload CSV/Excel", self._upload)
        ctk.CTkButton(uf, text="📋  Download Template", fg_color=OCEAN, hover_color=TEAL,
                      text_color=WHITE, width=160,
                      command=lambda: download_template(
                          ["port_name","country","arrival_date","cruise_id"],
                          "ports_template.csv"
                      )).pack(side="left", padx=3)

        self.table = DataTable(self, ["ID","Port Name","Country","Arrival Date","Cruise ID"])
        self.table.pack(fill="both", expand=True, padx=20, pady=(0,20))
        self.refresh()

    def refresh(self):
        conn = get_connection(); cursor = conn.cursor()
        cursor.execute("SELECT * FROM Ports")
        self.table.load(cursor.fetchall())
        cursor.close(); conn.close()

    def _cruise_ids(self):
        conn = get_connection(); cursor = conn.cursor()
        cursor.execute("SELECT cruise_id FROM Cruises")
        ids = [str(r[0]) for r in cursor.fetchall()]
        cursor.close(); conn.close(); return ids

    def _add(self):
        ids = self._cruise_ids()
        fields = [
            ("Port Name",    "e.g. Nassau",  None),
            ("Country",      "e.g. Bahamas", None),
            ("Arrival Date", "YYYY-MM-DD",   None),
            ("Cruise ID",    f"One of: {', '.join(ids)}", None),
        ]
        def submit(v, err):
            ok,pn  = validate_text(v["Port Name"],"Port Name");
            if not ok: err.configure(text=pn);  return False
            ok,co  = validate_text(v["Country"],  "Country");
            if not ok: err.configure(text=co);  return False
            ok,ad  = validate_date(v["Arrival Date"],"Arrival Date");
            if not ok: err.configure(text=ad);  return False
            ok,cid = validate_int(v["Cruise ID"],"Cruise ID");
            if not ok: err.configure(text=cid); return False
            conn = get_connection(); cursor = conn.cursor()
            cursor.execute("SELECT cruise_id FROM Cruises WHERE cruise_id=%s",(cid,))
            if not cursor.fetchone(): err.configure(text=f"Cruise ID {cid} not found."); cursor.close(); conn.close(); return False
            cursor.execute("INSERT INTO Ports (port_name,country,arrival_date,cruise_id) VALUES (%s,%s,%s,%s)",(pn,co,ad,cid))
            conn.commit(); cursor.close(); conn.close(); self.refresh(); return True
        FormDialog(self, "➕  Add Port", fields, submit)

    def _edit(self):
        row = self.table.selected_row()
        if not row: messagebox.showwarning("⚠️ No Selection","Please select a port first."); return
        fields = [("Port Name","",None),("Country","",None),("Arrival Date","YYYY-MM-DD",None),("Cruise ID","",None)]
        prefill = {"Port Name":str(row[1]),"Country":str(row[2]),
                   "Arrival Date":str(row[3]),"Cruise ID":str(row[4])}
        pid = row[0]
        def submit(v, err):
            ok,pn  = validate_text(v["Port Name"],"Port Name");
            if not ok: err.configure(text=pn);  return False
            ok,co  = validate_text(v["Country"],  "Country");
            if not ok: err.configure(text=co);  return False
            ok,ad  = validate_date(v["Arrival Date"],"Arrival Date");
            if not ok: err.configure(text=ad);  return False
            ok,cid = validate_int(v["Cruise ID"],"Cruise ID");
            if not ok: err.configure(text=cid); return False
            conn = get_connection(); cursor = conn.cursor()
            cursor.execute("SELECT cruise_id FROM Cruises WHERE cruise_id=%s",(cid,))
            if not cursor.fetchone(): err.configure(text=f"Cruise ID {cid} not found."); cursor.close(); conn.close(); return False
            cursor.execute("UPDATE Ports SET port_name=%s,country=%s,arrival_date=%s,cruise_id=%s WHERE port_id=%s",
                           (pn,co,ad,cid,pid))
            conn.commit(); cursor.close(); conn.close(); self.refresh(); return True
        FormDialog(self, "✏️  Edit Port", fields, submit, prefill)

    def _delete(self):
        pid = self.table.selected_id()
        if not pid: messagebox.showwarning("⚠️ No Selection","Please select a port first."); return
        if messagebox.askyesno("🗑️ Confirm Delete",f"Delete Port {pid}?"):
            conn = get_connection(); cursor = conn.cursor()
            cursor.execute("DELETE FROM Ports WHERE port_id=%s",(pid,)); conn.commit()
            cursor.close(); conn.close()
        self.refresh()

    def _upload(self):
        path = filedialog.askopenfilename(filetypes=[("CSV/Excel","*.csv *.xlsx")])
        if not path: return
        try: rows = read_file(path)
        except Exception as e: messagebox.showerror("❌ File Error", str(e)); return
        required = {"port_name","country","arrival_date","cruise_id"}
        if not required.issubset(set(rows[0].keys()) if rows else set()):
            messagebox.showerror("❌ Wrong Format",f"File must have columns:\n{', '.join(required)}"); return
        success=0; errors=[]
        conn = get_connection(); cursor = conn.cursor()
        for i, row in enumerate(rows, 2):
            try:
                ok,pn  = validate_text(row.get("port_name",""),   "port_name")
                ok2,co = validate_text(row.get("country",""),     "country")
                ok3,ad = validate_date(row.get("arrival_date",""),"arrival_date")
                ok4,cid= validate_int(row.get("cruise_id",""),    "cruise_id")
                if not all([ok,ok2,ok3,ok4]): errors.append(f"Row {i}: validation failed"); continue
                cursor.execute("SELECT cruise_id FROM Cruises WHERE cruise_id=%s",(cid,))
                if not cursor.fetchone(): errors.append(f"Row {i}: Cruise ID {cid} not found"); continue
                cursor.execute("INSERT INTO Ports (port_name,country,arrival_date,cruise_id) VALUES (%s,%s,%s,%s)",
                               (pn,co,ad,cid)); success+=1
            except Exception as e: errors.append(f"Row {i}: {e}")
        conn.commit(); cursor.close(); conn.close()
        self.refresh()
        UploadResultDialog(self, "📂 Port Upload Results", success, errors)


# ─────────────────────────────────────────
#  PRODUCTS TAB
# ─────────────────────────────────────────
class ProductsTab(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent, fg_color=NAVY)
        tab_header(self, "📦  Products & Activities", [
            ("➕  Add",    TEAL, SEAFOAM, NAVY, self._add),
            ("✏️  Edit",   OCEAN, TEAL, WHITE,  self._edit),
            ("🗑️  Delete", RED, "#b91c1c", WHITE, self._delete),
        ], self.refresh)

        uf = ctk.CTkFrame(self, fg_color="transparent")
        uf.pack(fill="x", padx=20, pady=(0,8))
        upload_btn(uf, "📂  Upload CSV/Excel", self._upload)
        ctk.CTkButton(uf, text="📋  Download Template", fg_color=OCEAN, hover_color=TEAL,
                      text_color=WHITE, width=160,
                      command=lambda: download_template(
                          ["name","category","price"],
                          "products_template.csv"
                      )).pack(side="left", padx=3)
        ctk.CTkLabel(uf, text=f"  Categories: {', '.join(CATEGORIES)}",
                     font=("Helvetica",11), text_color=GRAY).pack(side="left", padx=8)

        self.table = DataTable(self, ["ID","Name","Category","Price ($)"])
        self.table.pack(fill="both", expand=True, padx=20, pady=(0,20))
        self.refresh()

    def refresh(self):
        conn = get_connection(); cursor = conn.cursor()
        cursor.execute("SELECT product_id,name,category,price FROM Products")
        self.table.load(cursor.fetchall())
        cursor.close(); conn.close()

    def _add(self):
        fields = [
            ("Product Name","e.g. Margarita, Snorkeling Tour, Hot Stone Massage", None),
            ("Category","Select category", CATEGORIES),
            ("Price","e.g. 12.99", None),
        ]
        def submit(v, err):
            ok,name  = validate_text(v["Product Name"],"Product Name");
            if not ok: err.configure(text=name);  return False
            cat      = v["Category"]
            ok,price = validate_price(v["Price"]);
            if not ok: err.configure(text=price); return False
            conn = get_connection(); cursor = conn.cursor()
            cursor.execute("INSERT INTO Products (name,category,price) VALUES (%s,%s,%s)",(name,cat,price))
            conn.commit(); cursor.close(); conn.close(); self.refresh(); return True
        FormDialog(self, "➕  Add Product / Activity", fields, submit)

    def _edit(self):
        row = self.table.selected_row()
        if not row: messagebox.showwarning("⚠️ No Selection","Please select a product first."); return
        pid = row[0]
        fields = [("Product Name","",None),("Category","",CATEGORIES),("Price","",None)]
        prefill = {"Product Name":str(row[1]),"Category":str(row[2]),"Price":str(row[3])}
        def submit(v, err):
            ok,name  = validate_text(v["Product Name"],"Product Name");
            if not ok: err.configure(text=name);  return False
            cat      = v["Category"]
            ok,price = validate_price(v["Price"]);
            if not ok: err.configure(text=price); return False
            conn = get_connection(); cursor = conn.cursor()
            cursor.execute("UPDATE Products SET name=%s,category=%s,price=%s WHERE product_id=%s",(name,cat,price,pid))
            conn.commit(); cursor.close(); conn.close(); self.refresh(); return True
        FormDialog(self, "✏️  Edit Product", fields, submit, prefill)

    def _delete(self):
        pid = self.table.selected_id()
        if not pid: messagebox.showwarning("⚠️ No Selection","Please select a product first."); return
        conn = get_connection(); cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM Purchases WHERE product_id=%s",(pid,))
        pc = cursor.fetchone()[0]
        if pc > 0:
            messagebox.showerror("🚫 Cannot Delete",f"Product {pid} appears in {pc} purchase(s).\nDelete those first.")
            cursor.close(); conn.close(); return
        if messagebox.askyesno("🗑️ Confirm Delete",f"Delete Product {pid}?"):
            cursor.execute("DELETE FROM Products WHERE product_id=%s",(pid,)); conn.commit()
        cursor.close(); conn.close(); self.refresh()

    def _upload(self):
        path = filedialog.askopenfilename(filetypes=[("CSV/Excel","*.csv *.xlsx")])
        if not path: return
        try: rows = read_file(path)
        except Exception as e: messagebox.showerror("❌ File Error", str(e)); return
        required = {"name","category","price"}
        if not required.issubset(set(rows[0].keys()) if rows else set()):
            messagebox.showerror("❌ Wrong Format",f"File must have columns:\n{', '.join(required)}"); return
        success=0; errors=[]
        conn = get_connection(); cursor = conn.cursor()
        for i, row in enumerate(rows, 2):
            try:
                ok,name  = validate_text(row.get("name",""),    "name")
                cat      = row.get("category","").strip()
                ok2,price= validate_price(row.get("price",""))
                if not ok: errors.append(f"Row {i}: {name}"); continue
                if cat not in CATEGORIES: errors.append(f"Row {i}: category must be one of {CATEGORIES}"); continue
                if not ok2: errors.append(f"Row {i}: {price}"); continue
                cursor.execute("INSERT INTO Products (name,category,price) VALUES (%s,%s,%s)",(name,cat,price)); success+=1
            except Exception as e: errors.append(f"Row {i}: {e}")
        conn.commit(); cursor.close(); conn.close()
        self.refresh()
        UploadResultDialog(self, "📂 Product Upload Results", success, errors)


# ─────────────────────────────────────────
#  PURCHASES TAB
# ─────────────────────────────────────────
class PurchasesTab(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent, fg_color=NAVY)
        tab_header(self, "🛒  Purchases", [
            ("➕  Add",    TEAL, SEAFOAM, NAVY, self._add),
            ("🗑️  Delete", RED, "#b91c1c", WHITE, self._delete),
        ], self.refresh)

        uf = ctk.CTkFrame(self, fg_color="transparent")
        uf.pack(fill="x", padx=20, pady=(0,8))
        upload_btn(uf, "📂  Upload CSV/Excel", self._upload)
        ctk.CTkButton(uf, text="📋  Download Template", fg_color=OCEAN, hover_color=TEAL,
                      text_color=WHITE, width=160,
                      command=lambda: download_template(
                          ["passenger_id","product_id","quantity","purchase_date"],
                          "purchases_template.csv"
                      )).pack(side="left", padx=3)

        self.table = DataTable(self, ["ID","Passenger","Product","Category","Qty","Date","Total ($)"])
        self.table.pack(fill="both", expand=True, padx=20, pady=(0,20))
        self.refresh()

    def refresh(self):
        conn = get_connection(); cursor = conn.cursor()
        cursor.execute("""
                       SELECT pu.purchase_id,
                              CONCAT(pa.first_name,' ',pa.last_name),
                              pr.name, pr.category,
                              pu.quantity, pu.purchase_date,
                              ROUND(pu.quantity * pr.price, 2)
                       FROM Purchases pu
                                JOIN Passengers pa ON pu.passenger_id=pa.passenger_id
                                JOIN Products   pr ON pu.product_id=pr.product_id
                       ORDER BY pu.purchase_date DESC
                       """)
        self.table.load(cursor.fetchall())
        cursor.close(); conn.close()

    def _pax_list(self):
        conn = get_connection(); cursor = conn.cursor()
        cursor.execute("SELECT passenger_id, CONCAT(first_name,' ',last_name) FROM Passengers")
        rows = cursor.fetchall(); cursor.close(); conn.close()
        return [f"{r[0]} - {r[1]}" for r in rows]

    def _prod_list(self):
        conn = get_connection(); cursor = conn.cursor()
        cursor.execute("SELECT product_id, name, category FROM Products")
        rows = cursor.fetchall(); cursor.close(); conn.close()
        return [f"{r[0]} - {r[1]} ({r[2]})" for r in rows]

    def _add(self):
        pax = self._pax_list(); prods = self._prod_list()
        if not pax:   messagebox.showwarning("⚠️ No Passengers","Add a passenger first."); return
        if not prods: messagebox.showwarning("⚠️ No Products","Add products first."); return
        fields = [
            ("Passenger",     "Select passenger", pax),
            ("Product",       "Select product",   prods),
            ("Quantity",      "e.g. 1",           None),
            ("Purchase Date", "YYYY-MM-DD",       None),
        ]
        def submit(v, err):
            try:    pid   = int(v["Passenger"].split("-")[0].strip())
            except: err.configure(text="⚠️ Invalid passenger."); return False
            try:    prodid = int(v["Product"].split("-")[0].strip())
            except: err.configure(text="⚠️ Invalid product."); return False
            ok,qty = validate_int(v["Quantity"],"Quantity",1,999)
            if not ok: err.configure(text=qty); return False
            ok,dt  = validate_date(v["Purchase Date"],"Purchase Date")
            if not ok: err.configure(text=dt); return False
            conn = get_connection(); cursor = conn.cursor()
            cursor.execute("INSERT INTO Purchases (passenger_id,product_id,quantity,purchase_date) VALUES (%s,%s,%s,%s)",
                           (pid,prodid,qty,dt))
            conn.commit(); cursor.close(); conn.close(); self.refresh(); return True
        FormDialog(self, "🛒  Log Purchase", fields, submit)

    def _delete(self):
        pid = self.table.selected_id()
        if not pid: messagebox.showwarning("⚠️ No Selection","Please select a purchase first."); return
        if messagebox.askyesno("🗑️ Confirm Delete",f"Delete Purchase {pid}?"):
            conn = get_connection(); cursor = conn.cursor()
            cursor.execute("DELETE FROM Purchases WHERE purchase_id=%s",(pid,)); conn.commit()
            cursor.close(); conn.close()
        self.refresh()

    def _upload(self):
        path = filedialog.askopenfilename(filetypes=[("CSV/Excel","*.csv *.xlsx")])
        if not path: return
        try: rows = read_file(path)
        except Exception as e: messagebox.showerror("❌ File Error", str(e)); return
        required = {"passenger_id","product_id","quantity","purchase_date"}
        if not required.issubset(set(rows[0].keys()) if rows else set()):
            messagebox.showerror("❌ Wrong Format",f"File must have columns:\n{', '.join(required)}"); return
        success=0; errors=[]
        conn = get_connection(); cursor = conn.cursor()
        for i, row in enumerate(rows, 2):
            try:
                ok,pid   = validate_int(row.get("passenger_id",""),  "passenger_id")
                ok2,prid = validate_int(row.get("product_id",""),     "product_id")
                ok3,qty  = validate_int(row.get("quantity",""),       "quantity",1,999)
                ok4,dt   = validate_date(row.get("purchase_date",""), "purchase_date")
                if not all([ok,ok2,ok3,ok4]): errors.append(f"Row {i}: validation failed"); continue
                cursor.execute("SELECT passenger_id FROM Passengers WHERE passenger_id=%s",(pid,))
                if not cursor.fetchone(): errors.append(f"Row {i}: Passenger {pid} not found"); continue
                cursor.execute("SELECT product_id FROM Products WHERE product_id=%s",(prid,))
                if not cursor.fetchone(): errors.append(f"Row {i}: Product {prid} not found"); continue
                cursor.execute("INSERT INTO Purchases (passenger_id,product_id,quantity,purchase_date) VALUES (%s,%s,%s,%s)",
                               (pid,prid,qty,dt)); success+=1
            except Exception as e: errors.append(f"Row {i}: {e}")
        conn.commit(); cursor.close(); conn.close()
        self.refresh()
        UploadResultDialog(self, "📂 Purchase Upload Results", success, errors)


# ─────────────────────────────────────────
#  ANALYTICS TAB
# ─────────────────────────────────────────
class AnalyticsTab(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent, fg_color=NAVY)
        hdr = ctk.CTkFrame(self, fg_color="transparent")
        hdr.pack(fill="x", padx=20, pady=(20,10))
        ctk.CTkLabel(hdr, text="📊  Analytics Dashboard",
                     font=("Helvetica",22,"bold"), text_color=SEAFOAM).pack(side="left")
        ctk.CTkButton(hdr, text="↻  Refresh", fg_color=OCEAN, hover_color=TEAL,
                      text_color=WHITE, width=100, command=self.refresh).pack(side="right")

        ff = ctk.CTkFrame(self, fg_color=OCEAN, corner_radius=8)
        ff.pack(fill="x", padx=20, pady=(0,10))
        ctk.CTkLabel(ff, text="🔍  Filter by category:", font=("Helvetica",12),
                     text_color=WHITE).pack(side="left", padx=12, pady=8)
        self.cat_var = tk.StringVar(value="All")
        ctk.CTkOptionMenu(ff, values=["All"]+CATEGORIES, variable=self.cat_var,
                          fg_color=TEAL, button_color=OCEAN, text_color=WHITE,
                          command=lambda _: self.refresh()).pack(side="left", padx=8, pady=8)

        self.stats_frame  = ctk.CTkFrame(self, fg_color="transparent")
        self.stats_frame.pack(fill="x", padx=20, pady=(0,10))
        self.charts_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.charts_frame.pack(fill="both", expand=True, padx=20, pady=(0,20))
        self.refresh()

    def _stat_card(self, parent, icon, label, value, color=SEAFOAM):
        card = ctk.CTkFrame(parent, fg_color=OCEAN, corner_radius=10)
        card.pack(side="left", expand=True, fill="both", padx=6)
        ctk.CTkLabel(card, text=icon, font=("Helvetica",22)).pack(pady=(12,0))
        ctk.CTkLabel(card, text=str(value), font=("Helvetica",24,"bold"),
                     text_color=color).pack(pady=(2,0))
        ctk.CTkLabel(card, text=label, font=("Helvetica",11),
                     text_color=GRAY).pack(pady=(0,12))

    def refresh(self):
        for w in self.stats_frame.winfo_children():  w.destroy()
        for w in self.charts_frame.winfo_children(): w.destroy()

        cat    = self.cat_var.get()
        where  = "WHERE pr.category=%s" if cat != "All" else ""
        params = (cat,) if cat != "All" else ()

        conn = get_connection(); cursor = conn.cursor()

        cursor.execute(f"SELECT COALESCE(SUM(pu.quantity*pr.price),0) FROM Purchases pu JOIN Products pr ON pu.product_id=pr.product_id {where}", params)
        total_rev = float(cursor.fetchone()[0])

        cursor.execute(f"SELECT COALESCE(SUM(pu.quantity),0) FROM Purchases pu JOIN Products pr ON pu.product_id=pr.product_id {where}", params)
        total_qty = cursor.fetchone()[0]

        cursor.execute(f"SELECT COUNT(DISTINCT pu.passenger_id) FROM Purchases pu JOIN Products pr ON pu.product_id=pr.product_id {where}", params)
        unique_pax = cursor.fetchone()[0]

        cursor.execute(f"SELECT pr.name, SUM(pu.quantity) FROM Purchases pu JOIN Products pr ON pu.product_id=pr.product_id {where} GROUP BY pr.name ORDER BY 2 DESC LIMIT 1", params)
        top = cursor.fetchone()
        top_name = top[0][:16] if top else "N/A"

        self._stat_card(self.stats_frame, "💰", "Total Revenue",   f"${total_rev:,.2f}", GOLD)
        self._stat_card(self.stats_frame, "📦", "Items Sold",      total_qty,            SEAFOAM)
        self._stat_card(self.stats_frame, "👤", "Active Spenders", unique_pax,           GREEN)
        self._stat_card(self.stats_frame, "🏆", "Top Item",        top_name,             WHITE)

        cursor.execute(f"SELECT pr.name, SUM(pu.quantity) FROM Purchases pu JOIN Products pr ON pu.product_id=pr.product_id {where} GROUP BY pr.name ORDER BY 2 DESC LIMIT 10", params)
        top10 = cursor.fetchall()

        cursor.execute("SELECT pr.category, ROUND(SUM(pu.quantity*pr.price),2) FROM Purchases pu JOIN Products pr ON pu.product_id=pr.product_id GROUP BY pr.category ORDER BY 2 DESC")
        cat_rev = cursor.fetchall()

        cursor.close(); conn.close()

        plt.style.use("dark_background")
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(11,4))
        fig.patch.set_facecolor(NAVY)

        if top10:
            names = [r[0][:22] for r in top10]
            qtys  = [r[1] for r in top10]
            clrs  = [SEAFOAM,TEAL,GOLD,GREEN,"#a78bfa"]*2
            bars  = ax1.barh(names[::-1], qtys[::-1], color=clrs[:len(names)])
            ax1.set_facecolor(NAVY)
            ax1.set_title("🏆  Most Purchased Items", color=WHITE, fontsize=13, pad=12)
            ax1.tick_params(colors=WHITE, labelsize=9)
            ax1.spines[:].set_color(OCEAN)
            for bar, val in zip(bars, qtys[::-1]):
                ax1.text(bar.get_width()+0.05, bar.get_y()+bar.get_height()/2,
                         str(val), va="center", color=WHITE, fontsize=9)
        else:
            ax1.text(0.5,0.5,"No purchase data yet\nAdd products and log purchases!",
                     ha="center", va="center", color=GRAY, fontsize=12, transform=ax1.transAxes)
            ax1.set_facecolor(NAVY); ax1.set_title("Most Purchased Items", color=WHITE)

        if cat_rev:
            clabels = [r[0] for r in cat_rev]
            cvals   = [float(r[1]) for r in cat_rev]
            ax2.pie(cvals, labels=clabels, autopct="%1.1f%%",
                    colors=[SEAFOAM,TEAL,GOLD,GREEN,"#a78bfa"][:len(clabels)],
                    textprops={"color":WHITE,"fontsize":10})
            ax2.set_title("💰  Revenue by Category", color=WHITE, fontsize=13, pad=12)
        else:
            ax2.text(0.5,0.5,"No purchase data yet",
                     ha="center", va="center", color=GRAY, fontsize=12, transform=ax2.transAxes)
            ax2.set_title("Revenue by Category", color=WHITE)

        fig.tight_layout(pad=2)
        canvas = FigureCanvasTkAgg(fig, master=self.charts_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)
        plt.close(fig)


# ─────────────────────────────────────────
#  MAIN APP
# ─────────────────────────────────────────
class CruiseApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("🚢  Cruise Trip Tracker")
        self.geometry("1200x720")
        self.minsize(1000, 620)
        self.configure(fg_color=NAVY)
        setup_new_tables()

        if not cruise_exists():
            GateScreen(self, self._launch)
        else:
            self._launch()

    def _launch(self):
        for w in self.winfo_children(): w.destroy()
        self._build()

    def _build(self):
        sidebar = ctk.CTkFrame(self, width=210, fg_color=OCEAN, corner_radius=0)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)

        logo = ctk.CTkFrame(sidebar, fg_color=TEAL, height=90, corner_radius=0)
        logo.pack(fill="x"); logo.pack_propagate(False)
        ctk.CTkLabel(logo, text="🚢  Cruise Tracker",
                     font=("Helvetica",15,"bold"), text_color=WHITE).pack(expand=True)

        ctk.CTkLabel(sidebar, text="NAVIGATION", font=("Helvetica",10),
                     text_color=GRAY).pack(pady=(20,6), padx=16, anchor="w")

        self.nav_buttons = []
        nav = [("🚢  Cruises","Cruises"), ("👤  Passengers","Passengers"),
               ("⚓  Ports","Ports"),     ("📦  Products","Products"),
               ("🛒  Purchases","Purchases"), ("📊  Analytics","Analytics")]
        for label, key in nav:
            btn = ctk.CTkButton(sidebar, text=label,
                                fg_color="transparent", hover_color=TEAL,
                                text_color=WHITE, anchor="w", height=44,
                                font=("Helvetica",13),
                                command=lambda k=key: self._switch(k))
            btn.pack(fill="x", padx=8, pady=2)
            self.nav_buttons.append((key, btn))

        ctk.CTkLabel(sidebar, text="Database-backed App\nRalph Alexandre",
                     font=("Helvetica",10), text_color=GRAY).pack(side="bottom", pady=16)

        self.content = ctk.CTkFrame(self, fg_color=NAVY, corner_radius=0)
        self.content.pack(side="left", fill="both", expand=True)

        self.tabs = {
            "Cruises":    CruisesTab(self.content),
            "Passengers": PassengersTab(self.content),
            "Ports":      PortsTab(self.content),
            "Products":   ProductsTab(self.content),
            "Purchases":  PurchasesTab(self.content),
            "Analytics":  AnalyticsTab(self.content),
        }
        self._switch("Cruises")

    def _switch(self, name):
        for tab in self.tabs.values(): tab.pack_forget()
        self.tabs[name].pack(fill="both", expand=True)
        self.tabs[name].refresh()
        for key, btn in self.nav_buttons:
            btn.configure(fg_color=TEAL if key == name else "transparent")


if __name__ == "__main__":
    app = CruiseApp()
    app.mainloop()